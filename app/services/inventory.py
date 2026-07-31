from __future__ import annotations

import json
from collections import Counter
from datetime import datetime
from pathlib import Path

from app.services.importer import FIELD_ALIASES, REQUIRED_FIELDS


def _resolve_headers(headers: list[str]) -> dict[str, str]:
    resolved: dict[str, str] = {}
    values = {header.strip(): header for header in headers}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if alias in values:
                resolved[field] = values[alias]
                break
    return resolved


def inventory_workbooks(root: Path) -> dict[str, object]:
    from openpyxl import load_workbook

    files: list[dict[str, object]] = []
    header_patterns: Counter[tuple[str, ...]] = Counter()
    skipped: list[dict[str, str]] = []
    total_rows = 0
    for path in sorted(root.rglob("*.xlsx")):
        if ".downloading" in path.name:
            skipped.append({"path": str(path), "reason": "incomplete-download"})
            continue
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            first = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            headers = [str(value).strip() if value is not None else "" for value in first]
            mapping = _resolve_headers(headers)
            rows = max(int(sheet.max_row or 1) - 1, 0)
            workbook.close()
            relative = path.relative_to(root)
            parts = relative.parts
            files.append(
                {
                    "path": str(relative),
                    "category": parts[0] if len(parts) >= 3 else "",
                    "city_hint": parts[-2] if len(parts) >= 2 else path.stem,
                    "rows": rows,
                    "bytes": path.stat().st_size,
                    "headers": headers,
                    "mapping": mapping,
                    "missing_required": [
                        field for field in REQUIRED_FIELDS if field not in mapping
                    ],
                }
            )
            header_patterns[tuple(headers)] += 1
            total_rows += rows
        except Exception as exc:
            skipped.append({"path": str(path), "reason": str(exc)[:300]})
    return {
        "schema_version": "source-inventory-v1",
        "root": str(root.resolve()),
        "generated_at": datetime.now().isoformat(),
        "file_count": len(files),
        "estimated_rows": total_rows,
        "total_bytes": sum(int(item["bytes"]) for item in files),
        "unique_header_patterns": len(header_patterns),
        "files_missing_required": sum(bool(item["missing_required"]) for item in files),
        "files": files,
        "skipped": skipped,
    }


def write_inventory(root: Path, destination: Path) -> dict[str, object]:
    inventory = inventory_workbooks(root)
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_suffix(destination.suffix + ".tmp")
    temporary.write_text(json.dumps(inventory, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(destination)
    return inventory
